import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import ColorScaleRule

# ---------- Load all results ----------
df = pd.read_csv("./transactions_rule_scored.csv", parse_dates=["datetime"])
model_comparison = pd.read_csv("./model_comparison.csv")
lr_importance = pd.read_csv("./logreg_feature_importance.csv")
rf_importance = pd.read_csv("./rf_feature_importance.csv")
test_scored = pd.read_csv("./test_set_scored.csv", parse_dates=["datetime"])
rule_metrics = pd.read_csv("./rule_based_metrics.csv")

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1B1F3B", end_color="1B1F3B")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=16, color="1B1F3B")
SUB_FONT = Font(name=FONT, italic=True, size=10, color="666666")
KPI_LABEL_FONT = Font(name=FONT, size=10, color="666666")
KPI_VALUE_FONT = Font(name=FONT, bold=True, size=18, color="00A651")
NORMAL_FONT = Font(name=FONT, size=10)
BOLD_FONT = Font(name=FONT, bold=True, size=10)
RED_FONT = Font(name=FONT, bold=True, size=10, color="C00000")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def style_header(ws, row):
    for cell in ws[row]:
        if cell.value is not None:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_df(ws, data, start_row=1, as_table=False, table_name=None):
    for j, col_name in enumerate(data.columns, 1):
        ws.cell(row=start_row, column=j, value=col_name)
    for i, row in enumerate(data.itertuples(index=False), start_row + 1):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
    style_header(ws, start_row)
    last_row = start_row + len(data)
    last_col = get_column_letter(len(data.columns))
    if as_table and table_name:
        tbl = Table(displayName=table_name, ref=f"A{start_row}:{last_col}{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)
    return last_row

# ================= Sheet 1: Overview =================
ws1 = wb.active
ws1.title = "Overview"
ws1["A1"] = "Fraud & Risk Scoring Model"
ws1["A1"].font = TITLE_FONT
ws1["A2"] = "OPay-style Transaction Fraud Detection | Rule-Based vs Machine Learning Comparison"
ws1["A2"].font = SUB_FONT

total_txns = len(df)
total_fraud = int(df["is_fraud"].sum())
fraud_rate = total_fraud / total_txns

kpi_defs = [
    ("Total Transactions Analyzed", f"{total_txns:,}"),
    ("Confirmed Fraud Cases", f"{total_fraud}"),
    ("Fraud Rate", f"{fraud_rate*100:.2f}%"),
    ("Features Engineered", "7"),
]
for i, (label, val) in enumerate(kpi_defs):
    c = 1 + i * 2
    r = 4
    cl = get_column_letter(c)
    ws1.merge_cells(f"{cl}{r}:{get_column_letter(c+1)}{r}")
    ws1[f"{cl}{r}"] = label
    ws1[f"{cl}{r}"].font = KPI_LABEL_FONT
    ws1.merge_cells(f"{cl}{r+1}:{get_column_letter(c+1)}{r+1}")
    ws1[f"{cl}{r+1}"] = val
    ws1[f"{cl}{r+1}"].font = KPI_VALUE_FONT
    for rr in (r, r+1):
        for cc in (c, c+1):
            ws1.cell(row=rr, column=cc).fill = PatternFill("solid", start_color="F2F2F2")

ws1["A8"] = "Approach"
ws1["A8"].font = BOLD_FONT
approach_text = [
    "1. Engineered 7 fraud-risk features per transaction: amount deviation from customer's own history "
    "(z-score), new/unrecognized device, region mismatch vs home region, odd-hour timing, transaction "
    "velocity (1h/24h), and account age.",
    "2. Built a rule-based point-scoring system (auditable, explainable to non-technical stakeholders) "
    "as a baseline.",
    "3. Trained Logistic Regression and Random Forest models on the same features, using a time-based "
    "train/test split (train on earlier data, test on later data) to avoid lookahead bias.",
    "4. Compared precision, recall, F1, and ROC-AUC across all three approaches on the same held-out test period.",
]
r = 9
for line in approach_text:
    ws1.cell(row=r, column=1, value=line).font = NORMAL_FONT
    ws1.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws1.merge_cells(f"A{r}:H{r}")
    ws1.row_dimensions[r].height = 30
    r += 1
autosize(ws1, [16]*8)

# ================= Sheet 2: Model_Comparison =================
ws2 = wb.create_sheet("Model_Comparison")
ws2["A1"] = "Model Comparison — Rule-Based vs Logistic Regression vs Random Forest"
ws2["A1"].font = TITLE_FONT
ws2["A2"] = "All metrics evaluated on the same held-out test period (last ~25% of transactions chronologically)"
ws2["A2"].font = SUB_FONT
last2 = write_df(ws2, model_comparison, start_row=4, as_table=True, table_name="ModelComparison")
autosize(ws2, [22, 14, 12, 10, 10, 10])

bar = BarChart()
bar.title = "Precision, Recall, F1 by Model"
data = Reference(ws2, min_col=3, max_col=5, min_row=4, max_row=last2)
cats = Reference(ws2, min_col=1, min_row=5, max_row=last2)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.height, bar.width = 10, 20
ws2.add_chart(bar, f"A{last2 + 3}")

try:
    img_roc = Image("./roc_curve.png")
    img_roc.width, img_roc.height = 420, 350
    ws2.add_image(img_roc, f"H{last2 + 3}")
except Exception as e:
    print("ROC image embed skipped:", e)

try:
    img_pr = Image("./precision_recall_curve.png")
    img_pr.width, img_pr.height = 420, 350
    ws2.add_image(img_pr, f"H{last2 + 22}")
except Exception as e:
    print("PR image embed skipped:", e)

# ================= Sheet 3: Rule_Based_Scoring =================
ws3 = wb.create_sheet("Rule_Based_Scoring")
ws3["A1"] = "Rule-Based Scoring Methodology"
ws3["A1"].font = TITLE_FONT
rule_desc = [
    ["Rule", "Trigger Condition", "Points"],
    ["Amount anomaly", "Transaction amount > 3 std devs above customer's own historical average", 25],
    ["New device", "Device not previously associated with this customer", 30],
    ["Odd hour", "Transaction between 12:00 AM and 5:00 AM", 15],
    ["Velocity spike", "2 or more transactions by same customer within the last 1 hour", 30],
    ["Region mismatch", "Transaction region differs from customer's home region", 20],
]
r = 3
for row in rule_desc:
    for j, val in enumerate(row, 1):
        ws3.cell(row=r, column=j, value=val)
    r += 1
style_header(ws3, 3)
for row in ws3.iter_rows(min_row=4, max_row=r-1, max_col=3):
    for cell in row:
        cell.font = NORMAL_FONT
        cell.border = BORDER
autosize(ws3, [20, 60, 10])

tier_row = r + 2
ws3.cell(row=tier_row, column=1, value="Risk Tiers").font = BOLD_FONT
tiers = [["Tier", "Score Range", "Action"],
         ["High", "70+", "Auto-flag for manual review / temporary hold"],
         ["Medium", "40-69", "Secondary verification (OTP, device confirmation)"],
         ["Low", "15-39", "Monitor, no immediate action"],
         ["Minimal", "0-14", "No action"]]
r2 = tier_row + 1
for row in tiers:
    for j, val in enumerate(row, 1):
        ws3.cell(row=r2, column=j, value=val)
    r2 += 1
style_header(ws3, tier_row + 1)
for row in ws3.iter_rows(min_row=tier_row+2, max_row=r2-1, max_col=3):
    for cell in row:
        cell.font = NORMAL_FONT
        cell.border = BORDER

metrics_row = r2 + 2
ws3.cell(row=metrics_row, column=1, value="Rule-Based Performance (full dataset, threshold = score >= 50)").font = BOLD_FONT
write_df(ws3, rule_metrics, start_row=metrics_row + 1, as_table=True, table_name="RuleMetrics")

# ================= Sheet 4: Feature_Importance =================
ws4 = wb.create_sheet("Feature_Importance")
ws4["A1"] = "Feature Importance — Logistic Regression Coefficients"
ws4["A1"].font = TITLE_FONT
ws4["A2"] = "Standardized coefficients: positive = increases fraud probability, negative = decreases it"
ws4["A2"].font = SUB_FONT
last4a = write_df(ws4, lr_importance, start_row=4, as_table=True, table_name="LRImportance")
autosize(ws4, [22, 16])

bar2 = BarChart()
bar2.title = "Logistic Regression Coefficients"
data = Reference(ws4, min_col=2, min_row=4, max_row=last4a)
cats = Reference(ws4, min_col=1, min_row=5, max_row=last4a)
bar2.add_data(data, titles_from_data=True)
bar2.set_categories(cats)
bar2.height, bar2.width = 9, 18
ws4.add_chart(bar2, f"A{last4a + 3}")

rf_start = last4a + 20
ws4.cell(row=rf_start, column=1, value="Feature Importance — Random Forest").font = TITLE_FONT
last4b = write_df(ws4, rf_importance, start_row=rf_start + 2, as_table=True, table_name="RFImportance")

bar3 = BarChart()
bar3.title = "Random Forest Feature Importance"
data = Reference(ws4, min_col=2, min_row=rf_start+2, max_row=last4b)
cats = Reference(ws4, min_col=1, min_row=rf_start+3, max_row=last4b)
bar3.add_data(data, titles_from_data=True)
bar3.set_categories(cats)
bar3.height, bar3.width = 9, 18
ws4.add_chart(bar3, f"A{last4b + 3}")

# ================= Sheet 5: Flagged_Transactions =================
ws5 = wb.create_sheet("Flagged_Transactions")
ws5["A1"] = "High-Risk Flagged Transactions (Test Period)"
ws5["A1"].font = TITLE_FONT
ws5["A2"] = "Transactions where the ML model's fraud probability exceeded 0.5, sorted by risk"
ws5["A2"].font = SUB_FONT
flagged = test_scored[test_scored["fraud_probability_rf"] >= 0.5].copy()
flagged = flagged.sort_values("fraud_probability_rf", ascending=False)
display_cols = ["transaction_id", "customer_id", "datetime", "amount", "channel",
                "transaction_region", "amount_zscore", "is_new_device", "region_mismatch",
                "txn_count_last_1h", "rule_score", "fraud_probability_rf", "is_fraud"]
last5 = write_df(ws5, flagged[display_cols], start_row=4, as_table=True, table_name="FlaggedTxns")
autosize(ws5, [14, 14, 18, 12, 12, 14, 12, 10, 12, 10, 10, 16, 10])
rule = ColorScaleRule(start_type="min", start_color="FFFFFFFF",
                      end_type="max", end_color="FFC00000")
last_col_letter = get_column_letter(len(display_cols))
ws5.conditional_formatting.add(f"{get_column_letter(display_cols.index('fraud_probability_rf')+1)}5:"
                                f"{get_column_letter(display_cols.index('fraud_probability_rf')+1)}{last5}", rule)

# ================= Sheet 6: Data_Dictionary =================
ws6 = wb.create_sheet("Data_Dictionary")
ws6["A1"] = "Data Dictionary & Methodology"
ws6["A1"].font = TITLE_FONT
dict_rows = [
    ["Field", "Description", "Type"],
    ["transaction_id", "Unique transaction identifier", "Text"],
    ["customer_id", "Unique customer identifier", "Text"],
    ["datetime", "Transaction timestamp", "Datetime"],
    ["amount", "Transaction amount (NGN)", "Numeric"],
    ["device_id", "Device used for the transaction", "Text"],
    ["transaction_region", "Region where transaction occurred", "Text"],
    ["is_fraud", "Ground truth label (1 = confirmed fraud, 0 = legitimate)", "Binary"],
    ["amount_zscore", "How many std devs this amount is from the customer's own prior average", "Numeric (engineered)"],
    ["is_new_device", "1 if device differs from customer's known primary device", "Binary (engineered)"],
    ["region_mismatch", "1 if transaction region differs from customer's home region", "Binary (engineered)"],
    ["is_odd_hour", "1 if transaction occurred between 12am-5am", "Binary (engineered)"],
    ["txn_count_last_1h", "Number of this customer's transactions in the preceding 1 hour", "Numeric (engineered)"],
    ["txn_count_last_24h", "Number of this customer's transactions in the preceding 24 hours", "Numeric (engineered)"],
    ["account_age_days", "Customer's account age in days", "Numeric"],
    ["rule_score", "Sum of rule-based risk points", "Numeric (engineered)"],
    ["fraud_probability_rf", "Random Forest model's predicted fraud probability", "Numeric (model output)"],
]
r = 3
for row in dict_rows:
    for j, val in enumerate(row, 1):
        ws6.cell(row=r, column=j, value=val)
    r += 1
style_header(ws6, 3)
for row in ws6.iter_rows(min_row=4, max_row=r-1, max_col=3):
    for cell in row:
        cell.font = NORMAL_FONT
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
autosize(ws6, [22, 60, 20])

method_row = r + 2
ws6.cell(row=method_row, column=1, value="Methodology").font = BOLD_FONT
methodology = [
    "1. Synthetic dataset: 600 customers, ~38,600 transactions over 9 months, with a realistic 0.79% fraud rate.",
    "2. Fraud simulated as account-takeover bursts (new device, region change, high velocity, odd hours) PLUS "
    "a harder 'subtle fraud' category (1-2 transactions, familiar device, elevated but not extreme amount) "
    "to avoid an unrealistically easy detection problem.",
    "3. Legitimate 'anomaly' transactions were deliberately included (real travel, real device upgrades, real "
    "large purchases) so the model has to distinguish true fraud from superficially similar legitimate behavior.",
    "4. Time-based train/test split (not random) to simulate real deployment: train on the past, predict the future.",
    "5. Three approaches compared on the identical test period: rule-based scoring, Logistic Regression, Random Forest.",
]
for i, line in enumerate(methodology):
    c = ws6.cell(row=method_row+1+i, column=1, value=line)
    c.font = NORMAL_FONT
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws6.merge_cells(f"A{method_row+1+i}:D{method_row+1+i}")
    ws6.row_dimensions[method_row+1+i].height = 30

# ================= Sheet 7: Insights_Recommendations =================
ws7 = wb.create_sheet("Insights_Recommendations")
ws7["A1"] = "Insights & Recommendations"
ws7["A1"].font = TITLE_FONT
ws7.column_dimensions["A"].width = 4
ws7.column_dimensions["B"].width = 100

rule_row = model_comparison[model_comparison["model"] == "Rule-Based"].iloc[0]
lr_row = model_comparison[model_comparison["model"] == "Logistic Regression"].iloc[0]
rf_row = model_comparison[model_comparison["model"] == "Random Forest"].iloc[0]

insights = [
    ("Finding", f"The rule-based system achieves very high precision ({rule_row['precision']:.3f}) but lower "
                f"recall ({rule_row['recall']:.3f}) than the ML models on the same test period -- it reliably "
                "catches obvious account-takeover bursts but misses subtler fraud (1-2 transactions on a "
                "familiar device, moderately elevated amount)."),
    ("Recommendation", "Use the rule-based system as a fast, explainable first-line filter for obvious cases "
                        "(useful for real-time blocking), but do not rely on it alone -- pair it with a "
                        "learned model to catch subtler patterns."),
    ("Finding", f"Both ML models achieve recall of 1.000 on the test period (catching every confirmed fraud "
                f"case) but at the cost of more false positives than the rule-based approach -- Random Forest "
                f"flags {rf_row['precision']:.3f} precision vs the rule-based system's {rule_row['precision']:.3f}."),
    ("Recommendation", "This is a genuine precision-recall trade-off, not a flaw: in fraud detection, missing "
                        "a fraud case (false negative) is typically far more costly than a false positive "
                        "(a legitimate customer facing extra verification). Tune the decision threshold based "
                        "on the actual cost ratio the business assigns to each error type."),
    ("Finding", "Amount z-score, new-device flag, and region mismatch are consistently the top 3 predictive "
                "features across both Logistic Regression and Random Forest -- transaction velocity matters "
                "but less than expected once the other three signals are present."),
    ("Recommendation", "Prioritize real-time device fingerprinting and geolocation checks in the production "
                        "fraud stack, since these two signals alone carry most of the model's discriminating power."),
    ("Finding", "Model performance here (ROC-AUC > 0.999) is unusually strong because the synthetic fraud "
                "signal was deliberately made learnable; real transaction data has more overlap between fraud "
                "and legitimate behavior, and true performance would be lower."),
    ("Recommendation", "Before production deployment, validate against a labeled sample of real transactions "
                        "and expect precision/recall to be meaningfully lower than shown here -- this project "
                        "demonstrates the method and pipeline, not a production-ready accuracy claim."),
]
r = 3
for label, text in insights:
    ws7[f"A{r}"] = label
    ws7[f"A{r}"].font = BOLD_FONT if label == "Finding" else RED_FONT
    ws7[f"B{r}"] = text
    ws7[f"B{r}"].font = NORMAL_FONT
    ws7[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws7.row_dimensions[r].height = 60
    r += 2

wb.save("./OPay_Fraud_Risk_Scoring_Model.xlsx")
print("Saved OPay_Fraud_Risk_Scoring_Model.xlsx")
